import requests
import openpyxl
from bs4 import BeautifulSoup

url = "https://tw.stock.yahoo.com/quote/" # 奇摩股市url
targetfile = '投資總表.xlsx'              # 目標檔案
targetsheet = '庫存股票'                  # 目標工作表名稱
targetnumber = ''                        # 目標股號
column = 1                               # 工作表目標 欄
targetrow = 3                            # 工作表目標 列
pricecol = 7                             # 目標股價 欄
headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36'
}

# 開檔案
workbook = openpyxl.load_workbook(targetfile)
# 抓到所有資料表
sheet = workbook.worksheets
# 抓到資料表的個數
sheetlength = len(sheet)
# 搜尋所有資料表的名字假如與targetsheet相同就跳出去執行搜尋股價的程式
for targetsheetnumber in range(0,sheetlength):
    # print(targetsheetnumber)
    if sheet[targetsheetnumber].title==targetsheet:
        break
print(" 股號 ","股價")
# 台股971支不會超過這個range
for row in range(targetrow,900):
    # 如果為空單位即停止搜尋
    if sheet[targetsheetnumber].cell(row,column).value==None:
        break
    
    # 強制字串化
    targetnumber = str(sheet[targetsheetnumber].cell(row,column).value)

    # strip()去除前後空白 split('ETF')分割資料為純str(數字) 並且只留下股號的部分
    targetnumber = targetnumber.strip().split('ETF')
    targetnumber = targetnumber[len(targetnumber)-1]

    # 要求資料(url , headers , timeout) headers timeout逾期時長為選填
    r = requests.get(str(url)+str(targetnumber), timeout=5)

    # 如果成功要求到資料則執行
    if r.status_code == 200:
        # 整理
        soup = BeautifulSoup(r.text,"html.parser").select("#atomic .Fz\(32px\)")
        # 取出數字輸出
        for price in soup:
            price = price.text
        print(targetnumber , price)
        # 目標股價儲存格更改
        sheet[targetsheetnumber].cell(row,pricecol).value=float(price)
    else:
        print(targetnumber , "未讀取到資料")
        sheet[targetsheetnumber].cell(row,pricecol).value="-"
# 儲存
workbook.save(targetfile)