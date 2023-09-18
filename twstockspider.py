#! python3
#==============================================================================
#    Copyright (c) 2023 Jimmy Li. All rights reserved.
#    This program contains proprietary and confidential information.
#    All rights reserved except as may be permitted by prior written consent.
#
#    ModuleName:
#            twstockspider.py
#
#    Abstract:
#            search target stock from yahoostock
#    Author:
#            10-Feb-2023 Jimmy Li
#
#    Revision History:
#           Rev  1.0.0.2 18-Sep-2023 Jimmy Li
#                   1.create logger class
#                   2.add try except
#           Rev  1.0.0.1 10-Feb-2023 Jimmy Li
#                First create.
#==============================================================================

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font
from bs4 import BeautifulSoup
import os
import sys
from datetime import datetime
import logging
from logging.handlers import TimedRotatingFileHandler

class config():
    __version__ = "1.0.0.2"
    LogFileName = datetime.now().strftime("logs/twstockspider_%Y-%m-%d.log")
    url = "https://tw.stock.yahoo.com/quote/"               # 奇摩股市url
    targetFileName = os.path.join(os.getcwd(), '投資總表.xlsx')  # 目標檔案
    targetSheetTitle = '庫存股票'                                 # 目標工作表名稱
    targetnumber = ''                                       # 目標股號
    stockNumberColumn = 1                                              # 工作表目標 欄
    dataStartRow = 3                            # 工作表目標 列
    priceColumn = 7                             # 目標股價 欄
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36'
    }
    bRes = True
    targetSheetIndex = ''

class Logger():
    def __init__(self, LogFileName):

        self.logger = logging.getLogger(LogFileName)
        self.logger.setLevel(logging.DEBUG)

        self.fileHandler = TimedRotatingFileHandler(LogFileName, when = 'midnight', backupCount = 31, encoding = 'utf-8')
        self.fileHandler.setLevel(logging.DEBUG)

        self.consoleHandler = logging.StreamHandler()
        self.consoleHandler.setLevel(logging.DEBUG)

        logFileFormate = logging.Formatter(
            '[%(asctime)s][%(levelname)5s] - %(funcName)10s - %(lineno)-4d - %(message)s',
            datefmt = '%Y/%m/%d %H:%M:%S')
        consoleFormatter = logging.Formatter(
            '[%(levelname)5s] - %(funcName)10s - %(lineno)-4d - %(message)s')
        
        self.fileHandler.setFormatter(logFileFormate)
        self.consoleHandler.setFormatter(consoleFormatter)

        if not len(self.logger.handlers):
            self.logger.addHandler(self.fileHandler)
            self.logger.addHandler(self.consoleHandler)

    def getlog(self):
        return self.logger
    
    def closeLog(self):
        if len(self.logger.handlers):
            self.logger.removeHandler(self.fileHandler)
            self.logger.removeHandler(self.consoleHandler)
            self.fileHandler.flush()
            self.fileHandler.close()
            self.consoleHandler.flush()
            self.consoleHandler.close()

if __name__ == "__main__":

    setting = config()
    if not os.path.exists('logs'):
        os.makedirs('logs')

    logModule = Logger(setting.LogFileName)
    g_logger = logModule.getlog()
    g_logger.info(f"股市搜集器 {setting.__version__}")

    try:
        # 開檔案
        workbook = openpyxl.load_workbook(setting.targetFileName)
        # 抓到所有資料表
        sheet = workbook.worksheets
        # 抓到資料表的個數
        sheetlength = len(sheet)
        # 搜尋所有資料表的名字假如與targetsheet相同就跳出去執行搜尋股價的程式
        for sheetIndex in range(0,sheetlength):
            # print(sheetIndex)
            if sheet[sheetIndex].title == setting.targetSheetTitle:
                setting.targetSheetIndex = sheetIndex
                break
    except:
        strFailReason = "Exception Fail (%)" % (str(sys.exc_info()[1]))
        g_logger.error(strFailReason)
        setting.bRes = False

    if setting.bRes:
        g_logger.info("股號 , 股價")
        # 台股971支不會超過這個range
        for row in range(setting.dataStartRow, 900):
            try:
                # 如果為空單位即停止搜尋
                if sheet[setting.targetSheetIndex].cell(row, setting.stockNumberColumn).value == None:
                    break
                
                # 強制字串化
                targetnumber = str(sheet[setting.targetSheetIndex].cell(row, setting.stockNumberColumn).value)

                # strip()去除前後空白 split('ETF')分割資料為純str(數字) 並且只留下股號的部分
                targetnumber = targetnumber.strip().split('ETF')
                targetnumber = targetnumber[len(targetnumber)-1]

                # 要求資料(url , headers , timeout) headers timeout逾期時長為選填
                requestHtml = requests.get(str(setting.url) + str(targetnumber), timeout = 5)

                # 如果成功要求到資料則執行
                if requestHtml.status_code == 200:
                    # 整理
                    soup = BeautifulSoup(requestHtml.text, "html.parser").select("#atomic .Fz\(32px\)")
                    # 取出數字輸出
                    for Price in soup:
                        strPrice = Price.text
                    g_logger.info(f"{targetnumber} , {strPrice}")
                    try:
                        # 目標股價儲存格更改
                        sheet[setting.targetSheetIndex].cell(row, setting.priceColumn).value = float(strPrice)
                        sheet[setting.targetSheetIndex].cell(row, setting.priceColumn).fill = PatternFill()
                        sheet[setting.targetSheetIndex].cell(row, setting.priceColumn).font = Font(name = '微軟正黑體', size = 14, color = "000000", bold = True)
                    except:
                        sheet[setting.targetSheetIndex].cell(row, setting.priceColumn).fill = PatternFill(fgcolor = "ff0000", fill_type = "solid")
                        sheet[setting.targetSheetIndex].cell(row, setting.priceColumn).font = Font(name = '微軟正黑體', size = 14, color = "ffffff", bold = True)
                else:
                    g_logger.info(f"股票編號 {targetnumber} 未讀取到資料")
                    sheet[setting.targetSheetIndex].cell(row, setting.priceColumn).fill = PatternFill(fgcolor = "ff0000", fill_type = "solid")
                    sheet[setting.targetSheetIndex].cell(row, setting.priceColumn).font = Font(name = '微軟正黑體', size = 14, color = "ffffff", bold = True)
            except:
                strFailReason = "Exception Fail (%)" % (str(sys.exc_info()[1]))
                g_logger.error(strFailReason)
        # 儲存
        workbook.save(setting.targetFileName)
    else:
        strFailReason = "開啟excel檔案失敗或excel格式錯誤"
        g_logger.error(strFailReason)
    
    input("輸入任意鍵以結束程式...")